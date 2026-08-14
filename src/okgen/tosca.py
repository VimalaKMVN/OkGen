"""Run TOSCA Script — populate a TOSCA input workbook from selected files.

For the selected files, resolve each file's (Chain, Process, Format) into the
exact strings the TOSCA workbook expects (chain/process via config name maps;
Format via the workbook's own ``Key`` sheet, which is what the cell dropdowns
reference), dedupe to unique combinations, and write them into the chosen
workbook's data sheet — CONTIGUOUS from the top, clearing everything below so
TOSCA (which stops at the first blank row) processes exactly the selected set.

The workbooks are macro-enabled (.xlsm) with x14 dropdown validations that an
openpyxl round-trip would silently drop, so the write is a TARGETED edit of just
the data-sheet cells inside the .xlsm zip — macros, dropdowns, formatting and the
~180 helper columns are left byte-untouched (same philosophy as the fixed-width
byte-span edits). Reading (the Key sheet) uses openpyxl; only writing is surgical.

Works on BOTH engines: the Calgary JSON layouts and the line-based ``.OK`` ones.
Only two values are read per file — Chain and ticket Format — so the difference
is confined to ``_json_header`` vs ``_okfile_header``; everything downstream
(dedupe, Key-sheet resolution, the workbook write, the .bat launch) is
layout-agnostic. A ``.OK`` file and the JSON for the same combination therefore
collapse to ONE row, which is the point of deduping by (Chain, Process, Format).

Config: ``config/tosca.yaml`` (see that file).
"""

from __future__ import annotations

import datetime
import os
import re
import subprocess
import sys
import time
import zipfile
from pathlib import Path
from typing import List, Optional, Tuple
from xml.sax.saxutils import escape

from okgen import paths as fs
from okgen.detect import detect_layout


class ToscaError(Exception):
    """A configuration/workbook-level failure (not a per-file issue)."""


_WORKBOOK_EXTS = (".xlsm", ".xlsx")
_BAT_EXTS = (".bat", ".cmd")


def _resolve_one(raw: str, exts, kind: str) -> Path:
    """Resolve a config path that may be:
      * a full FILE path (used as-is), or
      * a GLOB pattern (e.g. ``…\\*ExecutionScript*.bat``) — the single match, or
      * a FOLDER holding one matching file (``.xlsm`` / ``.bat``).
    Raises ToscaError with a clear message if it's missing or ambiguous."""
    p = Path(raw)
    if p.is_file():
        return p

    def _pick(hits, where):
        hits = sorted(h for h in hits
                      if h.is_file() and h.suffix.lower() in exts
                      and not h.name.startswith("~$"))
        if len(hits) == 1:
            return hits[0]
        if not hits:
            raise ToscaError(f"no {kind} ({'/'.join(exts)}) {where}")
        raise ToscaError(f"multiple {kind} files {where} — narrow it with a full "
                         f"path or a name pattern: " + ", ".join(h.name for h in hits))

    if any(ch in p.name for ch in "*?[") and not p.is_dir():   # glob pattern
        parent = p.parent
        return _pick(parent.glob(p.name) if parent.is_dir() else [], f"matching {raw}")
    if p.is_dir():
        return _pick(p.iterdir(), f"in folder: {p}")
    raise ToscaError(f"{kind} path not found: {raw}")


# --------------------------------------------------------------------------- #
# Value resolution
# --------------------------------------------------------------------------- #
def _json_header(path: Path) -> dict:
    import json
    data = json.loads(fs.read_text(path, "utf-8")).get("data", {})
    return data.get("header", {}) or {}


def _okfile_header(path: Path, layout, config) -> dict:
    """``chain``/``format`` for a line-based (.OK) layout, read from its header
    record — the same two values ``_json_header`` yields for a JSON file.

    ``format`` is the file's TICKET format (the Key-sheet code). It is read from
    a real header field when the layout has one, else from a DERIVED field of
    that name (EUCartonLabel computes its format from distribution/pack type,
    see D8). The EU GTA ``process`` field is deliberately NOT used as a fallback:
    it is the layout discriminator (``D``->StyleHeader, ``H``->CartonLabel), not
    a ticket format, and 'H' collides with a real format code in other columns.
    """
    from okgen.okfile import parse_okfile

    okf = parse_okfile(path, layout)
    header = (okf.layout.sections or [None])[0]
    if header is None:
        return {}
    recs = okf.sections().get(header.name) or []
    if not recs:
        return {}
    rec = recs[0]
    values = {f.name: rec.get(f.name) for f in header.fields}

    out = {"chain": values.get("chain")}
    if "format" in values:
        out["format"] = values["format"]
    else:
        for spec in config.derived_fields(okf.layout.name):
            if spec.get("name") == "format":
                out["format"] = _format_code(config.eval_derived(spec, values))
                break
    return out


def format_head(value) -> Optional[str]:
    """The CODE at the head of a ``"code - description"`` string.

    Splits on the FIRST hyphen, not on ``" -"``, and that is deliberate: the
    workbooks' own Key sheets are not consistently spaced. Two rows are typed
    without the space before the hyphen — ``'T-Q-Line Small Gum Label'`` (TJMaxx
    SH/PT) and ``'J- Rat Tail Gum Label'`` (Marshalls SH/PT) — so a ``" -"``
    split returned the WHOLE string as the code and those two formats resolved
    to nothing at all: a file carrying format ``T`` was reported as *not in Key
    column C* and dropped from the run. A format code is always the leading
    token and never contains a hyphen, so the first hyphen is the right seam.

    The same rule names the ticket-format FOLDER (see ``_find_child``), so the
    Key row and the folder are identified by one definition rather than two.
    """
    if value in (None, ""):
        return value
    return str(value).split("-", 1)[0].strip()


# Derived fields carry the full ``"1 - Carton Label"`` label; the code is the
# same leading token, found the same way.
_format_code = format_head


def _format_string(key_ws, column: str, code: str) -> Optional[str]:
    """The exact ``"code - description"`` from a Key column whose value starts
    with ``code`` (e.g. Key!F for Winners Style Header, code 'B' -> 'B - Blue Gum')."""
    code = str(code).strip()
    for row in range(2, (key_ws.max_row or 2) + 1):
        v = key_ws[f"{column}{row}"].value
        if v is None:
            continue
        if format_head(v) == code:
            return str(v)
    return None


def build_rows(paths, registry, config, key_ws) -> Tuple[List[dict], List[dict]]:
    """(unique rows, per-file errors). Each row: chain/process/format/status/
    source/date. Deduped by (chain, process, format).

    Each row also carries ``files`` — the selected paths that resolved to it.
    Staging needs exactly that: the row IS the input folder, so the files
    sharing a row are the files that belong in one folder. It is dropped by the
    sheet write, which only ever reads the columns named in ``columns:``.
    """
    t = config.tosca()
    chain_names = t.get("chain_names", {}) or {}
    process_names = t.get("process_names", {}) or {}
    format_columns = t.get("format_columns", {}) or {}
    defaults = t.get("defaults", {}) or {}
    europe = set(t.get("europe_chain_names", []) or [])
    date_eu = t.get("date_format_europe", "%d/%m/%Y")
    date_def = t.get("date_format_default", "%m/%d/%Y")
    today = datetime.date.today()

    seen: dict = {}
    rows: List[dict] = []
    errors: List[dict] = []
    for p in paths or []:
        p = Path(p)
        name = p.name
        try:
            layout = detect_layout(p).layout
        except Exception as exc:                       # noqa: BLE001
            errors.append({"file": name, "error": f"could not read: {exc}"})
            continue
        reg_layout = registry.get(layout) if layout else None
        if reg_layout is None:
            errors.append({"file": name, "error": f"unknown layout {layout!r}"})
            continue

        try:
            header = (_json_header(p) if getattr(reg_layout, "json_mode", False)
                      else _okfile_header(p, reg_layout, config))
        except Exception as exc:                       # noqa: BLE001
            errors.append({"file": name, "error": f"could not read header: {exc}"})
            continue
        raw_chain = header.get("chain")
        info = config.chain(raw_chain)                 # resolves a code OR a name
        chain_code = info.code if info else None
        chain_name = chain_names.get(chain_code)
        process_name = process_names.get(layout)
        fmt_code = header.get("format")

        if not chain_name:
            errors.append({"file": name, "error": f"no TOSCA chain name for chain {raw_chain!r}"})
            continue
        if not process_name:
            errors.append({"file": name, "error": f"no TOSCA process name for layout {layout}"})
            continue
        column = (format_columns.get(chain_name) or {}).get(process_name)
        if not column:
            errors.append({"file": name,
                           "error": f"no Format column mapped for {chain_name} / {process_name}"})
            continue
        if fmt_code in (None, ""):
            errors.append({"file": name,
                           "error": f"{layout} has no ticket 'format' value to map "
                                    f"(no format field, and none derived)"})
            continue
        fmt_str = _format_string(key_ws, column, fmt_code)
        if fmt_str is None:
            errors.append({"file": name,
                           "error": f"format {fmt_code!r} not in Key column {column} "
                                    f"for {chain_name}/{process_name}"})
            continue

        key = (chain_name, process_name, fmt_str)
        if key in seen:
            # Dedupe unique combinations — but the FILE still belongs to the
            # row, because every file resolving to a combination is staged into
            # that combination's one folder.
            seen[key]["files"].append(str(p))
            continue
        is_eu = chain_name in europe
        row = {
            "chain": chain_name,
            "process": process_name,
            "format": fmt_str,
            "status": defaults.get("status", "Work Pending"),
            "source": defaults.get("source", "Online"),
            "date": today.strftime(date_eu if is_eu else date_def),
            "files": [str(p)],
        }
        seen[key] = row
        rows.append(row)
    return rows, errors


# --------------------------------------------------------------------------- #
# Input file staging — putting the selected files where TOSCA reads them
# --------------------------------------------------------------------------- #
# Updating the sheet only tells TOSCA WHICH combinations to process; it reads the
# files themselves from a tree of its own, addressed as
# ``{B[Chain]}\{B[Process]}\{B[Format]}`` — the very triple a row already is. So
# nothing new is resolved here: the ROW IS THE FOLDER, and the files that share a
# row are the files that belong in one folder.
#
# A run is therefore three ordered steps: stage, write the sheet, fire the .bat.
# Staging goes FIRST so that a folder-level failure leaves the workbook exactly
# as it was — there is nothing to undo, and the sheet can never end up listing a
# combination whose folder was not set up.
#
# Config: the ``input_staging`` block plus each script's ``input_folders``.

STAGE_OK = "ok"                    # folder found, named exactly as the sheet cell
STAGE_NAME_MISMATCH = "name_mismatch"   # found by CODE, but named differently
STAGE_MISSING = "missing"          # no folder for this format at all
STAGE_CREATE = "create"            # missing, and create_missing is on

# Files cleared from a target folder, by the engine whose files are going in.
_ENGINE_EXTS = {"ok": (".ok",), "json": (".json",)}


def staging_config(config) -> dict:
    """The ``input_staging`` block with defaults applied. Absent block = OFF, so
    a config written before this feature keeps behaving exactly as it did."""
    st = (config.tosca() or {}).get("input_staging") or {}
    return {
        "enabled": bool(st.get("enabled", False)),
        "subpath": str(st.get("subpath") or "{chain}\\{process}\\{format}"),
        "match_by_code": bool(st.get("match_format_by_code", True)),
        "clear": str(st.get("clear", "matching")).strip().lower(),
        "create_missing": bool(st.get("create_missing", False)),
        "overwrite": bool(st.get("overwrite", True)),
    }


def _script_roots(script: dict) -> List[str]:
    raw = script.get("input_folders") or []
    if isinstance(raw, str):
        raw = [raw]
    return [str(r).strip() for r in raw if str(r).strip()]


def _children(parent: Path) -> List[str]:
    """Sub-folder NAMES of ``parent`` ([] if it isn't a folder). Listed through
    ``long_path`` because these trees are deep enough to pass MAX_PATH (D44),
    and only the names are kept so the caller keeps building real paths."""
    try:
        return [c.name for c in Path(fs.long_path(parent)).iterdir() if c.is_dir()]
    except (OSError, ValueError):
        return []


def _find_child(parent: Path, name: str, by_code: bool = False):
    """Locate one folder level. Returns ``(path, how)`` with ``how`` one of
    ``exact`` / ``case`` / ``code``, or ``(None, None)``.

    ``code`` is the interesting one and applies to the ticket-format level only:
    it finds the folder whose leading code matches even when the full name does
    not, which is how the Key sheet's ``'T-Q-Line Small Gum Label'`` is
    recognised as the tree's ``'T - Q-Line Small Gum Label'``. That is a
    DETECTION, not a repair — see ``plan_staging``.
    """
    direct = parent / name
    try:
        if direct.is_dir():
            return direct, "exact"
    except OSError:
        pass
    kids = _children(parent)
    low = str(name).lower()
    for k in kids:                                   # Windows folders are
        if k.lower() == low:                         # case-insensitive; a
            return parent / k, "case"                # case-sensitive FS is not
    if by_code:
        code = format_head(name)
        if code not in (None, ""):
            for k in kids:
                if format_head(k) == code:
                    return parent / k, "code"
    return None, None


def _leaf_segments(subpath: str, row: dict) -> List[str]:
    """The configured ``{chain}\\{process}\\{format}`` filled in from the row.
    Split on BOTH separators so a config written with Windows backslashes also
    resolves on the POSIX box the tests run on."""
    filled = subpath.format(chain=row.get("chain", ""),
                            process=row.get("process", ""),
                            format=row.get("format", ""))
    return [seg for seg in re.split(r"[\\/]+", filled) if seg]


def _clear_list(folder: Path, exts, mode: str, keep: set) -> List[str]:
    """Files this run would delete from ``folder``.

    Only FILES, never sub-folders, and only in a folder this run is populating —
    a stale file under some other format is harmless, because the data sheet is
    cleared below the last row, so TOSCA processes only the combinations written.

    ``keep`` holds the sources being staged INTO this folder. A file already
    sitting in its own target (the user browsed the TOSCA tree in OkGen, edited,
    and ran) must not be deleted out from under itself — it is about to be
    rewritten from the snapshot anyway.
    """
    if mode == "none":
        return []
    out = []
    try:
        entries = sorted(Path(fs.long_path(folder)).iterdir(), key=lambda c: c.name)
    except (OSError, ValueError):
        return []
    for child in entries:
        if not child.is_file():
            continue
        real = folder / child.name
        if str(real).lower() in keep:
            continue
        if mode == "all" or child.suffix.lower() in exts:
            out.append(str(real))
    return out


def plan_staging(rows, script, config, engines=None) -> dict:
    """What staging WOULD do — pure inspection, nothing written.

    Backs both the confirmation dialog (so a destructive step is seen before it
    is agreed to) and the run itself, so the preview and the action cannot
    describe different work.

    A row whose folder is missing, or whose folder disagrees with the sheet
    cell, is EXCLUDED and reported; every other selected combination still runs.
    Excluding rather than staging-anyway is not caution, it is arithmetic: TOSCA
    builds its input path from the Format CELL, so a row whose cell does not name
    a real folder cannot be processed however the files are arranged.
    """
    cfg = staging_config(config)
    roots = _script_roots(script)
    plan = {"enabled": cfg["enabled"], "configured": bool(roots), "roots": roots,
            "targets": [], "excluded": [], "remove_total": 0, "copy_total": 0,
            "clear": cfg["clear"], "create_missing": cfg["create_missing"]}
    if not cfg["enabled"] or not roots:
        return plan

    exts = tuple(sorted({e for eng in (engines or ["ok", "json"])
                         for e in _ENGINE_EXTS.get(eng, ())}))
    for row in rows:
        files = [str(f) for f in (row.get("files") or [])]
        keep = {f.lower() for f in files}
        targets, problems = [], []

        # Two selected files with the same NAME resolving to the same
        # combination would land on top of each other in the one folder, and
        # whichever copied last would win silently. Reported, and NEITHER is
        # staged — a bulk write path must not quietly keep one of two files.
        counts: dict = {}
        for f in files:
            counts[Path(f).name.lower()] = counts.get(Path(f).name.lower(), 0) + 1
        clashes = sorted(n for n, c in counts.items() if c > 1)
        if clashes:
            problems.append(
                "two or more selected files share a name and resolve to this same "
                "folder, so one would overwrite the other: " + ", ".join(clashes))

        for root in ([] if problems else roots):
            segs = _leaf_segments(cfg["subpath"], row)
            here, ok = Path(root), True
            for i, seg in enumerate(segs):
                last = (i == len(segs) - 1)
                found, how = _find_child(here, seg, by_code=last and cfg["match_by_code"])
                if found is None:
                    if last and cfg["create_missing"]:
                        targets.append({"root": root, "path": str(here / seg),
                                        "status": STAGE_CREATE, "how": None,
                                        "note": f"folder does not exist yet and will be created: {here / seg}",
                                        "remove": [], "copy": files})
                    else:
                        problems.append(
                            f"no {'ticket format' if last else 'folder'} {seg!r} under {here}")
                    ok = False
                    break
                if last and how == "code" and found.name != seg:
                    # States BOTH spellings as facts and prescribes NEITHER.
                    #
                    # This used to end "...until the workbook's Key sheet is
                    # corrected to <folder name>", which assumed the folder was
                    # right and the sheet wrong. The user hit the opposite — a
                    # mistyped FOLDER beside a correct sheet — where that advice
                    # would have edited a correct workbook to match the typo,
                    # putting the error somewhere the next person would trust.
                    #
                    # OkGen can see the two disagree; it cannot know which is
                    # true. The GTA UI is what settles it, so the message points
                    # there instead of choosing. (Repairing neither is D79.)
                    problems.append(
                        f"the Key sheet says {seg!r} but the folder is named "
                        f"{found.name!r}. TOSCA builds its input path from the "
                        f"Key sheet. Fix the folder name and/or the Key sheet "
                        f"name to the correct format name, to match the GTA UI.")
                    ok = False
                    break
                here = found
            if not ok:
                continue
            targets.append({"root": root, "path": str(here), "status": STAGE_OK,
                            "how": "exact",
                            "remove": _clear_list(here, exts, cfg["clear"], keep),
                            "copy": files})

        if problems:
            plan["excluded"].append({
                "chain": row.get("chain"), "process": row.get("process"),
                "format": row.get("format"),
                "files": [Path(f).name for f in files],
                "reasons": problems,
            })
            continue
        for t in targets:
            t.update(chain=row.get("chain"), process=row.get("process"),
                     format=row.get("format"))
            plan["targets"].append(t)
            plan["remove_total"] += len(t["remove"])
            plan["copy_total"] += len(t["copy"])
    return plan


def included_rows(rows, plan) -> List[dict]:
    """The rows staging did not exclude. An excluded combination is left out of
    the SHEET as well: writing it would ask TOSCA to process a folder that is not
    there, and report a failure that looks like TOSCA's rather than the Key
    sheet's."""
    if not plan.get("enabled") or not plan.get("configured"):
        return list(rows)
    bad = {(e.get("chain"), e.get("process"), e.get("format"))
           for e in plan.get("excluded", [])}
    return [r for r in rows if (r.get("chain"), r.get("process"), r.get("format")) not in bad]


def apply_staging(plan: dict) -> dict:
    """Carry out ``plan``: snapshot, clear, copy. Raises ToscaError on any
    folder-level failure, having deleted nothing.

    SNAPSHOT FIRST, and that is the whole design. The obvious order — clear the
    folder, then copy the files in — destroys its own input when a selected file
    already lives in the folder being cleared, which is exactly what happens when
    someone browses the TOSCA input tree in OkGen, edits a file and runs. Copying
    every source into a temp folder before anything is deleted makes the order
    irrelevant, and means a read failure costs nothing: the delete only ever
    happens once the replacement is in hand.
    """
    import shutil
    import tempfile

    targets = plan.get("targets") or []
    result = {"folders": [], "removed": 0, "copied": 0, "created": 0}
    if not targets:
        return result

    tmp = Path(tempfile.mkdtemp(prefix="okgen-tosca-"))
    try:
        snap = {}                                     # source path -> temp copy
        for t in targets:
            for src in t["copy"]:
                if src in snap:
                    continue
                s = Path(src)
                if not s.is_file():
                    raise ToscaError(f"could not stage input files — this file is "
                                     f"no longer there: {src}")
                dst = tmp / f"{len(snap)}_{s.name}"
                try:
                    fs.copy2(s, dst)
                except OSError as exc:
                    raise ToscaError(f"could not read {s.name} to stage it: {exc}")
                snap[src] = dst

        for t in targets:
            folder = Path(t["path"])
            entry = {"path": t["path"], "chain": t.get("chain"),
                     "process": t.get("process"), "format": t.get("format"),
                     "created": False, "removed": [], "copied": []}
            if t["status"] == STAGE_CREATE:
                try:
                    fs.mkdir(folder, parents=True, exist_ok=True)
                except OSError as exc:
                    raise ToscaError(f"could not create the input folder {folder}: {exc}")
                entry["created"] = True
                result["created"] += 1
            for victim in t["remove"]:
                try:
                    fs.unlink(Path(victim))
                except OSError as exc:
                    raise ToscaError(
                        f"could not clear the input folder {folder} — {Path(victim).name} "
                        f"could not be removed ({exc}). Nothing has been written to the "
                        f"workbook and TOSCA was not started.")
                entry["removed"].append(Path(victim).name)
                result["removed"] += 1
            for src in t["copy"]:
                dest = folder / Path(src).name
                try:
                    fs.copy2(snap[src], dest)
                except OSError as exc:
                    raise ToscaError(f"could not copy {Path(src).name} into {folder}: {exc}")
                entry["copied"].append(Path(src).name)
                result["copied"] += 1
            result["folders"].append(entry)
    finally:
        shutil.rmtree(tmp, ignore_errors=True)
    return result


# --------------------------------------------------------------------------- #
# Targeted .xlsm data-sheet write (macros/dropdowns preserved)
# --------------------------------------------------------------------------- #
def _col_num(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def _cell_xml(ref: str, style_attr: str, value: Optional[str]) -> str:
    if value is None:
        return f'<c r="{ref}"{style_attr}/>'
    return (f'<c r="{ref}"{style_attr} t="inlineStr">'
            f'<is><t xml:space="preserve">{escape(value)}</t></is></c>')


def _set_cell(xml: str, ref: str, value: Optional[str]) -> str:
    """Set (or clear, value=None) one cell in a worksheet XML string, inserting
    the cell/row if absent. Only touches that cell — everything else is byte-exact."""
    col = re.match(r"[A-Z]+", ref).group(0)
    rownum = ref[len(col):]
    cell_pat = re.compile(r'<c r="' + re.escape(ref) + r'"(?:\s[^>]*?)?(?:/>|>.*?</c>)', re.S)
    m = cell_pat.search(xml)
    if m:
        s = re.search(r'\ss="\d+"', m.group(0))
        return xml[:m.start()] + _cell_xml(ref, s.group(0) if s else "", value) + xml[m.end():]
    if value is None:
        return xml                                      # already absent == empty
    new_cell = _cell_xml(ref, "", value)
    row_pat = re.compile(r'(<row r="' + rownum + r'"[^>]*>)(.*?)(</row>)', re.S)
    rm = row_pat.search(xml)
    if rm:                                              # row exists, cell doesn't -> insert in order
        body = rm.group(2)
        at = len(body)
        for cm in re.finditer(r'<c r="([A-Z]+)\d+"', body):
            if _col_num(cm.group(1)) > _col_num(col):
                at = cm.start()
                break
        return xml[:rm.start(2)] + body[:at] + new_cell + body[at:] + xml[rm.end(2):]
    # row missing -> insert a new <row> in row-number order
    new_row = f'<row r="{rownum}">{new_cell}</row>'
    sd = re.search(r'(<sheetData>)(.*?)(</sheetData>)', xml, re.S)
    body = sd.group(2)
    at = len(body)
    for rr in re.finditer(r'<row r="(\d+)"', body):
        if int(rr.group(1)) > int(rownum):
            at = rr.start()
            break
    return xml[:sd.start(2)] + body[:at] + new_row + body[at:] + xml[sd.end(2):]


def _sheet_xml_path(names, wb_xml: str, rels_xml: str, sheet_name: str) -> str:
    rid = None
    for sm in re.finditer(r"<sheet\b[^>]*/?>", wb_xml):
        if f'name="{sheet_name}"' in sm.group(0):
            r = re.search(r'r:id="([^"]+)"', sm.group(0))
            rid = r.group(1) if r else None
            break
    if rid is None:
        raise ToscaError(f"data sheet {sheet_name!r} not found in workbook")
    rm = re.search(r'<Relationship[^>]*Id="' + re.escape(rid) + r'"[^>]*Target="([^"]+)"', rels_xml)
    if rm is None:
        raise ToscaError(f"could not resolve sheet target for {sheet_name!r}")
    target = rm.group(1)
    return ("xl/" + target) if not target.startswith("/") else target[1:]


def write_data_sheet(workbook: Path, data_sheet: str, rows: List[dict],
                     first_data_row: int, columns: dict, max_clear_row: int) -> None:
    """Write ``rows`` into the data sheet contiguously from ``first_data_row`` and
    clear the field columns of every row below, editing only the sheet XML inside
    the .xlsm zip (all other parts — macros, validations, styles — preserved)."""
    workbook = Path(workbook)
    with zipfile.ZipFile(fs.long_path(workbook)) as zf:
        names = zf.namelist()
        data = {n: zf.read(n) for n in names}
    wb_xml = data["xl/workbook.xml"].decode("utf-8")
    rels_xml = data["xl/_rels/workbook.xml.rels"].decode("utf-8")
    sheet_path = _sheet_xml_path(names, wb_xml, rels_xml, data_sheet)
    xml = data[sheet_path].decode("utf-8")

    for i, row in enumerate(rows):
        rnum = first_data_row + i
        for field, col in columns.items():
            xml = _set_cell(xml, f"{col}{rnum}", row.get(field))
    for rnum in range(first_data_row + len(rows), max_clear_row + 1):
        for col in columns.values():
            xml = _set_cell(xml, f"{col}{rnum}", None)

    data[sheet_path] = xml.encode("utf-8")
    tmp = workbook.with_name(workbook.name + ".okgen.tmp")
    with zipfile.ZipFile(fs.long_path(tmp), "w", zipfile.ZIP_DEFLATED) as zf:
        for n in names:                                 # preserve entry order
            zf.writestr(n, data[n])
    # Atomically swap in the new workbook. On Windows this fails if the file is
    # open (Excel) or held by a running TOSCA — retry briefly for a transient
    # handle, then give up cleanly (caller turns it into a friendly message).
    for attempt in range(5):
        try:
            fs.replace(tmp, workbook)
            return
        except PermissionError:
            if attempt == 4:
                try:
                    fs.unlink(tmp)
                except OSError:
                    pass
                raise
            time.sleep(0.25)



# --------------------------------------------------------------------------- #
# Releasing a workbook left open by a stopped TOSCA run
# --------------------------------------------------------------------------- #
CLOSE_OK = "closed"             # we closed it; nothing was lost
CLOSE_UNSAVED = "unsaved"       # it has unsaved edits — deliberately NOT closed
CLOSE_NO_EXCEL = "no_excel"     # no Excel we can talk to in this session
CLOSE_NOT_OPEN = "not_open"     # Excel is running but does not have this file
CLOSE_UNSUPPORTED = "unsupported"   # not Windows
CLOSE_ERROR = "error"

# Ask Excel to close ONE workbook. Deliberately not "kill EXCEL.EXE": that would
# take down every other workbook the user has open, unsaved work included.
#
# `Saved` is the safety gate. If the workbook has unsaved changes we refuse and
# say so, because the likeliest unsaved change is the PowerForms link someone
# just set by hand (D21) — discarding that would be worse than the lock.
_CLOSE_PS = """
$ErrorActionPreference = 'Stop'
$target = '{path}'
try {{ $xl = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application') }}
catch {{ exit 2 }}
foreach ($wb in $xl.Workbooks) {{
  if ($wb.FullName -ieq $target) {{
    if ($wb.Saved) {{ $wb.Close($false); exit 0 }} else {{ exit 3 }}
  }}
}}
exit 4
"""

_CLOSE_EXITS = {0: CLOSE_OK, 2: CLOSE_NO_EXCEL, 3: CLOSE_UNSAVED, 4: CLOSE_NOT_OPEN}


def close_open_workbook(path: Path, timeout: float = 20.0) -> str:
    """Try to close ``path`` in a running Excel. Returns one of CLOSE_*.

    Only reaches Excel in the SAME interactive session — a TOSCA run under a
    different account, or an Excel that crashed hard enough to leave COM, is
    beyond it. Those come back as CLOSE_NO_EXCEL/CLOSE_NOT_OPEN and the caller
    explains rather than pretending.
    """
    if os.name != "nt":
        return CLOSE_UNSUPPORTED
    script = _CLOSE_PS.format(path=str(Path(path).resolve()).replace("'", "''"))
    try:
        proc = subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", script],
            capture_output=True, timeout=timeout,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
    except (OSError, subprocess.SubprocessError):
        return CLOSE_ERROR
    return _CLOSE_EXITS.get(proc.returncode, CLOSE_ERROR)


def _locked_message(workbook: Path, outcome: str) -> str:
    """Say what is actually wrong, and what will fix it."""
    name = workbook.name
    if outcome == CLOSE_UNSAVED:
        return (f"{name} is open in Excel WITH UNSAVED CHANGES, so OkGen did not "
                f"close it — that could have discarded work (the PowerForms link, "
                f"for instance). Save or discard it in Excel, then run again.")
    if outcome == CLOSE_OK:
        return (f"{name} was open in Excel; OkGen closed it, but the file is still "
                f"locked. Something else is holding it — most likely a TOSCA run "
                f"that is still going. Wait for it to finish, then run again.")
    if outcome in (CLOSE_NO_EXCEL, CLOSE_NOT_OPEN):
        return (f"could not update the workbook — it looks LOCKED: {name}. No Excel "
                f"in this session has it open, so the holder is probably a TOSCA run "
                f"that was stopped part-way. End that process (or sign out and back "
                f"in), then run again.")
    return (f"could not update the workbook — it looks LOCKED: {name}. Close it in "
            f"Excel (and make sure no earlier TOSCA run still has it open), then "
            f"run again.")


# --------------------------------------------------------------------------- #
# Launch the TOSCA .bat (fire-and-forget)
# --------------------------------------------------------------------------- #
def _launch_bat(bat: Path) -> None:
    """Start the .bat and return immediately (fire-and-forget — TOSCA runs long
    and does its own reporting)."""
    folder = str(bat.parent)
    if os.name == "nt":
        # Force a NEW, VISIBLE console window for the .bat so the user sees the
        # "TOSCA started" screen (and any error). CREATE_NEW_CONSOLE makes Windows
        # allocate a fresh console regardless of whether OkGen itself has one;
        # ``cmd /k`` keeps the window open so a quick failure doesn't vanish. cwd
        # = the bat's folder. Fire-and-forget. Falls back to the shell "open"
        # (like double-clicking) if that ever fails.
        CREATE_NEW_CONSOLE = 0x00000010
        try:
            subprocess.Popen(["cmd", "/k", str(bat)], cwd=folder,
                             creationflags=CREATE_NEW_CONSOLE, close_fds=True)
        except Exception:                               # noqa: BLE001
            os.startfile(str(bat))                      # last resort
    else:
        # Non-Windows (dev/CI): run the script directly in a new session. A real
        # .bat won't execute here, but an executable stub does — enough to prove
        # the launch path without a Windows box.
        subprocess.Popen([str(bat)], cwd=folder, start_new_session=True,
                         stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                         stdin=subprocess.DEVNULL, close_fds=True)


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #
def _script_engines(script: dict) -> set:
    """Engines a script accepts. Absent ``applies_to`` means BOTH, so a config
    written before the .OK/JSON split keeps working unchanged."""
    raw = script.get("applies_to") or ["ok", "json"]
    if isinstance(raw, str):
        raw = [raw]
    return {str(e).strip().lower() for e in raw}


def _engine_of(path: Path, registry) -> Optional[str]:
    """``'json'`` / ``'ok'`` for a file, or None when it can't be detected (those
    are left for build_rows to report as a normal per-file error)."""
    try:
        layout = detect_layout(path).layout
    except Exception:                                   # noqa: BLE001
        return None
    reg_layout = registry.get(layout) if layout else None
    if reg_layout is None:
        return None
    return "json" if getattr(reg_layout, "json_mode", False) else "ok"


def list_scripts(config) -> List[dict]:
    return [{"name": s.get("name"), "applies_to": sorted(_script_engines(s))}
            for s in config.tosca_scripts()]


def report_folders(config) -> List[dict]:
    """Scripts that declare a `results:` folder, for the TOSCA Reports picker.

    A script without the key is simply not offered — the other eight are a
    config edit away, no code. `exists` is checked so a folder that is not set
    up on THIS machine says so, instead of opening an empty window or failing
    silently: these paths live on the user's D: drive and only some are wired up.
    """
    out = []
    for s in config.tosca_scripts():
        folder = (s.get("results") or "").strip()
        if not folder:
            continue
        out.append({"name": s.get("name"), "folder": folder,
                    "exists": Path(folder).is_dir()})
    return out


def open_report_folder(config, script_name: str) -> dict:
    """Open one script's results folder in the OS file manager.

    Deliberately hands off to Explorer rather than listing the files in OkGen.
    The folders hold Word and Excel documents, which no browser can render
    without a heavy vendored converter (the D31 trade) — and Explorer already
    does everything wanted here: thumbnails, sorting, and a double-click that
    opens each document in its own application. Imitating it in a table would be
    more work and permanently worse.

    NOTE the path is passed RAW. `okgen.paths` prefixes `\\?\` for reads and
    writes past MAX_PATH (D44), but Explorer does not accept a prefixed path —
    the same reason `run()` hands `os.startfile` the plain .bat path.
    """
    script = next((s for s in config.tosca_scripts()
                   if s.get("name") == script_name), None)
    if script is None:
        raise ToscaError(f"no TOSCA script named {script_name!r}")
    folder = (script.get("results") or "").strip()
    if not folder:
        raise ToscaError(
            f"{script_name} declares no results folder — add a `results:` path "
            f"to that script in config/tosca.yaml and restart OkGen")
    if not Path(folder).is_dir():
        raise ToscaError(
            f"the results folder for {script_name} does not exist on this "
            f"machine:\n{folder}\n\nCheck the `results:` path in "
            f"config/tosca.yaml, or run the script once to create it.")
    _reveal(folder)
    return {"opened": folder, "script": script_name}


def _reveal(folder: str) -> None:
    """Show a folder in the OS file manager, IN FRONT of the browser.

    `os.startfile` opens Explorer but does not raise it: OkGen is a background
    process as far as Win32 is concerned, so the new window can land behind the
    browser — the same foreground-lock that made the folder chooser look like it
    had not opened (D24/v0.80.0). On Windows the folder is opened through the
    shell and then explicitly brought forward; anything that fails falls back to
    plain `os.startfile`, so the worst case is exactly today's behaviour.

    macOS and Linux need none of this — `open` and `xdg-open` activate the
    window themselves.
    """
    if os.name != "nt":
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.Popen([opener, folder], start_new_session=True)
        return
    try:
        _reveal_windows_front(folder)
    except Exception:                                  # noqa: BLE001
        os.startfile(folder)                           # noqa: S606


# Bringing a window to the FRONT is not the same as creating it: a background
# process cannot take foreground on its own, so we attach to the current
# foreground window's input queue for the hand-off and detach after. This is
# the same manoeuvre browse_folder uses for the chooser; it is repeated rather
# than shared because that code is confirmed working on the user's box and is
# not worth destabilising for a few lines of P/Invoke (see PLAN §6).
_REVEAL_PS = r"""
Add-Type -AssemblyName System.Windows.Forms
Add-Type @"
using System;
using System.Runtime.InteropServices;
public static class RevealFg {
  [DllImport("user32.dll")] public static extern bool SetForegroundWindow(IntPtr h);
  [DllImport("user32.dll")] public static extern bool BringWindowToTop(IntPtr h);
  [DllImport("user32.dll")] public static extern bool ShowWindow(IntPtr h, int n);
  [DllImport("user32.dll")] public static extern bool IsIconic(IntPtr h);
  [DllImport("user32.dll")] public static extern void keybd_event(byte k, byte s, uint f, UIntPtr e);
  [DllImport("user32.dll")] public static extern IntPtr GetForegroundWindow();
  [DllImport("user32.dll")] public static extern uint GetWindowThreadProcessId(IntPtr h, IntPtr pid);
  [DllImport("kernel32.dll")] public static extern uint GetCurrentThreadId();
  [DllImport("user32.dll")] public static extern bool AttachThreadInput(uint a, uint b, bool f);
}
"@
$target = $env:OKGEN_REVEAL_PATH
$shell = New-Object -ComObject Shell.Application

function Find-Window($path) {
  foreach ($w in $shell.Windows()) {
    try {
      if ($w.Document.Folder.Self.Path -eq $path) { return $w }
    } catch { }        # a non-folder shell window (IE) has no Document.Folder
  }
  return $null
}

# Reuse a window ALREADY showing this folder rather than stacking duplicates —
# clicking Reports twice should raise the one window, not open a second.
$win = Find-Window $target
if ($win -eq $null) {
  $shell.Open($target)
  for ($i = 0; $i -lt 25 -and $win -eq $null; $i++) {
    Start-Sleep -Milliseconds 120
    $win = Find-Window $target
  }
}
if ($win -ne $null) {
  $h = [IntPtr]$win.HWND
  if ([RevealFg]::IsIconic($h)) { [RevealFg]::ShowWindow($h, 9) | Out-Null }  # SW_RESTORE
  $fg = [RevealFg]::GetForegroundWindow()
  $fgTid = [RevealFg]::GetWindowThreadProcessId($fg, [IntPtr]::Zero)
  $myTid = [RevealFg]::GetCurrentThreadId()
  $attached = $false
  if ($fgTid -ne 0 -and $fgTid -ne $myTid) {
    $attached = [RevealFg]::AttachThreadInput($myTid, $fgTid, $true)
  }
  # Tapping Alt releases Windows' foreground lock for this thread.
  [RevealFg]::keybd_event(0x12, 0, 0, [UIntPtr]::Zero)
  [RevealFg]::keybd_event(0x12, 0, 2, [UIntPtr]::Zero)
  [RevealFg]::BringWindowToTop($h) | Out-Null
  [RevealFg]::SetForegroundWindow($h) | Out-Null
  if ($attached) { [RevealFg]::AttachThreadInput($myTid, $fgTid, $false) | Out-Null }
} else {
  # Never leave the user with nothing: if the window could not be located,
  # open it the ordinary way and let it land where it lands.
  Start-Process explorer.exe -ArgumentList $target
}
"""


def _reveal_windows_front(folder: str) -> None:
    """Open (or raise) an Explorer window for ``folder`` and bring it forward.

    The path is passed through the ENVIRONMENT, not interpolated into the
    script: a Windows path is full of backslashes and may contain quotes, and
    building PowerShell by string concatenation is how a path becomes code.
    """
    import base64

    env = dict(os.environ, OKGEN_REVEAL_PATH=folder)
    enc = base64.b64encode(_REVEAL_PS.encode("utf-16-le")).decode("ascii")
    subprocess.run(["powershell", "-NoProfile", "-STA", "-EncodedCommand", enc],
                   env=env, capture_output=True, text=True, timeout=30)


def scripts_for(paths, registry, config) -> dict:
    """The pickable scripts for a selection: each annotated with how many of the
    selected files it would actually run. The client offers the applicable ones
    so a user can't silently aim an .OK selection at a JSON workbook."""
    counts = {"ok": 0, "json": 0, None: 0}
    for p in paths or []:
        counts[_engine_of(Path(p), registry)] += 1
    out = []
    for s in config.tosca_scripts():
        engines = _script_engines(s)
        out.append({"name": s.get("name"),
                    "applies_to": sorted(engines),
                    "matches": sum(counts.get(e, 0) for e in engines)})
    return {"scripts": out, "counts": {"ok": counts.get("ok", 0),
                                       "json": counts.get("json", 0),
                                       "unknown": counts.get(None, 0)}}


def _prepare(paths, script_name, registry, config) -> dict:
    """Everything a run and a preview both need: the resolved script + workbook,
    the engine routing, and the rows the selection resolves to.

    Shared rather than duplicated so the preview and the run cannot describe
    different work — the whole point of showing a destructive step beforehand is
    that what is shown is what happens.
    """
    import openpyxl

    t = config.tosca()
    script = config.tosca_script(script_name)
    if script is None:
        raise ToscaError(f"no TOSCA script named {script_name!r} in config/tosca.yaml")
    workbook = _resolve_one(script.get("workbook", ""), _WORKBOOK_EXTS, "workbook")
    data_sheet = script["data_sheet"]
    key_sheet = script.get("key_sheet") or t.get("key_sheet", "Key")
    # Sheet layout falls back to the global block, but a script whose data sheet
    # is laid out differently can override it without touching code.
    first_data_row = int(script.get("first_data_row", t.get("first_data_row", 2)))
    columns = script.get("columns") or t.get("columns") or {}

    # Engine routing: .OK and JSON have separate workbooks/.bats, so a run only
    # takes the files its script accepts and REPORTS the rest as skipped rather
    # than silently dropping them (or aiming them at the wrong workbook).
    engines = _script_engines(script)
    use_paths, skipped = [], []
    for p in paths or []:
        eng = _engine_of(Path(p), registry)
        if eng is None or eng in engines:
            use_paths.append(p)                         # undetectable -> normal error
        else:
            skipped.append({"file": Path(p).name, "engine": eng,
                            "reason": f"{'JSON' if eng == 'json' else '.OK'} file — "
                                      f"not applicable to this script"})

    wb = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    try:
        if key_sheet not in wb.sheetnames:
            raise ToscaError(f"Key sheet {key_sheet!r} not in workbook")
        key_ws = wb[key_sheet]
        rows, errors = build_rows(use_paths, registry, config, key_ws)
        data_ws = wb[data_sheet] if data_sheet in wb.sheetnames else None
        max_row = (data_ws.max_row if data_ws else 0) or 0
    finally:
        wb.close()
    return {"script": script, "workbook": workbook, "data_sheet": data_sheet,
            "first_data_row": first_data_row, "columns": columns,
            "engines": engines, "rows": rows, "errors": errors,
            "skipped": skipped, "max_row": max_row, "t": t}


def preview(paths, script_name, registry, config) -> dict:
    """What a run WOULD do to the input folders — for the confirmation dialog.

    Deliberately its own round trip rather than part of the script list: the
    folders depend on the script CHOSEN, and scanning all ten scripts' trees to
    answer a question about one would be both slow and mostly wrong (most of
    those roots are not present on a given machine).
    """
    prep = _prepare(paths, script_name, registry, config)
    plan = plan_staging(prep["rows"], prep["script"], config, prep["engines"])
    pv = {
        "script": script_name,
        "enabled": plan["enabled"],
        "configured": plan["configured"],
        "rows": len(included_rows(prep["rows"], plan)),
        "targets": [{"path": t["path"], "chain": t.get("chain"),
                     "process": t.get("process"), "format": t.get("format"),
                     "status": t["status"],
                     "remove": [Path(v).name for v in t["remove"]],
                     "copy": [Path(c).name for c in t["copy"]]}
                    for t in plan["targets"]],
        "excluded": plan["excluded"],
        "remove_total": plan["remove_total"],
        "copy_total": plan["copy_total"],
        "errors": prep["errors"],
        "skipped": prep["skipped"],
    }
    # The per-combination roll-up the dialog renders, and the plan report text.
    # Both are built HERE so the dialog, its View report and the plan log all
    # come from one computation rather than three that can disagree.
    pv["combinations"] = plan_combinations(pv)
    pv["report"] = build_plan_report(pv, len(paths or []))
    return pv


# --------------------------------------------------------------------------- #
# The run report
#
# Deliberately the same shape as `nicelabel_post.build_report` / `_write_log`,
# because it is the same job: ONE function returns the text, the log file gets
# it and the window shows it. What a user pastes into a message is then exactly
# what the log says, by construction rather than by discipline.
# --------------------------------------------------------------------------- #

def default_log_folder() -> Path:
    """``logs`` next to OkGen itself — not in the user's data folders."""
    return Path(__file__).resolve().parents[2] / "logs"


def combinations(res: dict) -> List[dict]:
    """One entry per Chain/Process/Format, with what staging did to its folder.

    The run already knows both halves — a staged folder entry carries the same
    chain/process/format triple the row does, because the row IS the folder
    (that is the addressing rule staging is built on). Joining them HERE rather
    than in the UI is what lets the report and the summary window show the same
    numbers without either recomputing them.
    """
    st = res.get("staging") or {}
    per: dict = {}
    for f in st.get("folders") or []:
        key = (f.get("chain"), f.get("process"), f.get("format"))
        got = per.setdefault(key, {"copied": 0, "removed": 0, "created": 0,
                                   "paths": []})
        got["copied"] += len(f.get("copied") or [])
        got["removed"] += len(f.get("removed") or [])
        got["created"] += 1 if f.get("created") else 0
        got["paths"].append(f.get("path"))
    out = []
    for row in res.get("rows") or []:
        key = (row.get("chain"), row.get("process"), row.get("format"))
        got = per.get(key) or {}
        out.append({
            "chain": row.get("chain"), "process": row.get("process"),
            "format": row.get("format"), "status": "written",
            "copied": got.get("copied", 0), "removed": got.get("removed", 0),
            "created": got.get("created", 0), "paths": got.get("paths", []),
            "files": [Path(f).name for f in (row.get("files") or [])],
        })
    for x in st.get("excluded") or []:
        out.append({
            "chain": x.get("chain"), "process": x.get("process"),
            "format": x.get("format"), "status": "not_run",
            "copied": 0, "removed": 0, "created": 0, "paths": [],
            "files": list(x.get("files") or []),
            "reasons": list(x.get("reasons") or []),
        })
    return out


def _wrap_names(names, indent: int, width: int = 96) -> List[str]:
    """File names as wrapped, indented lines — one long comma run is unreadable
    in a log and unquotable in a message."""
    pad = " " * indent
    lines, cur = [], ""
    for i, n in enumerate(names):
        piece = n + ("," if i < len(names) - 1 else "")
        if cur and len(cur) + 1 + len(piece) > width:
            lines.append(pad + cur)
            cur = piece
        else:
            cur = (cur + " " + piece).strip()
    if cur:
        lines.append(pad + cur)
    return lines


def build_report(res: dict) -> str:
    """The run report, as plain text — the log file AND the View report window."""
    st = res.get("staging") or {}
    combos = combinations(res)
    not_run = [c for c in combos if c["status"] == "not_run"]
    lines = [
        "OkGen — Run TOSCA Script",
        f"Script   : {res.get('script')}"
        + (f"        (engine: {', '.join(res.get('applies_to') or [])})"
           if res.get("applies_to") else ""),
        f"Workbook : {res.get('workbook')}",
        f"Sheet    : {res.get('data_sheet')}",
        f"Started  : {res.get('started', '')}      "
        f"Elapsed: {res.get('elapsed_seconds', 0):.1f}s",
        f"Result   : {res.get('written', 0)} rows written, "
        f"{st.get('copied', 0)} files staged, {st.get('removed', 0)} removed, "
        f"{len(not_run)} combination(s) NOT run",
        "",
        "COMBINATIONS",
    ]
    if not combos:
        lines.append("  (none — nothing matched this script)")
    for c in combos:
        if c["status"] == "written":
            lines.append(
                f"  [WRITTEN ] {str(c['chain'] or ''):<12} {str(c['process'] or ''):<15}"
                f" {str(c['format'] or ''):<26} +{c['copied']}  -{c['removed']}"
                + ("  (folder created)" if c["created"] else ""))
            for p in c["paths"]:
                lines.append(f"             {p}")
            if c["removed"]:
                # the removed NAMES are not carried per combination; the folder
                # entries hold them, so they are printed from there below
                pass
            if c["files"]:
                lines.append("               copied :")
                lines += _wrap_names(c["files"], 24)
        else:
            lines.append(
                f"  [NOT RUN ] {str(c['chain'] or ''):<12} {str(c['process'] or ''):<15}"
                f" {str(c['format'] or ''):<26} {len(c['files'])} file(s)")
            for r in c.get("reasons") or []:
                lines.append(f"               reason : {r}")
            if c["files"]:
                lines.append("               files  :")
                lines += _wrap_names(c["files"], 24)
    # The removed names, per folder — the one thing a combination row cannot
    # carry, since two roots can stage the same combination into two folders.
    removed_any = [f for f in (st.get("folders") or []) if f.get("removed")]
    if removed_any:
        lines += ["", "FILES REMOVED BEFORE STAGING"]
        for f in removed_any:
            lines.append(f"  {f.get('path')}")
            lines += _wrap_names(list(f.get("removed") or []), 6)
    skipped = res.get("skipped") or []
    if skipped:
        lines += ["", f"NOT APPLICABLE TO THIS SCRIPT ({len(skipped)})"]
        for s in skipped:
            lines.append(f"  {str(s.get('file', '')):<34} {s.get('reason', '')}")
    errs = res.get("errors") or []
    if errs:
        lines += ["", f"COULD NOT BE USED ({len(errs)})"]
        for e in errs:
            lines.append(f"  {str(e.get('file', '')):<34} {e.get('error', '')}")
    if not st.get("enabled", False):
        lines += ["", "STAGING OFF — input_staging.enabled is false in config/tosca.yaml,"
                      "             so no files were copied into the TOSCA tree."]
    elif not st.get("configured", False):
        lines += ["", "NO INPUT FOLDERS — this script has no input_folders in "
                      "config/tosca.yaml,", "             so no files were copied."]
    lines += ["", "LAUNCH"]
    if res.get("launched"):
        lines.append(f"  started : {res.get('bat')}")
    elif res.get("launch_error"):
        lines.append(f"  NOT started : {res.get('launch_error')}")
    else:
        lines.append("  not started (no rows were written)")
    return "\n".join(lines) + "\n"


def plan_combinations(pv: dict) -> List[dict]:
    """One entry per Chain/Process/Format for the CONFIRMATION dialog.

    The same roll-up `combinations()` does for a finished run, from the plan
    instead of the outcome — so the dialog and the result window count the same
    way and a user comparing them is not comparing two different arithmetics.
    Two roots can stage one combination into two folders, hence the merge.
    """
    per: dict = {}
    order: List[tuple] = []
    for t in pv.get("targets") or []:
        key = (t.get("chain"), t.get("process"), t.get("format"))
        if key not in per:
            per[key] = {"chain": key[0], "process": key[1], "format": key[2],
                        "status": "will_run", "copy": 0, "remove": 0,
                        "create": 0, "paths": []}
            order.append(key)
        got = per[key]
        got["copy"] += len(t.get("copy") or [])
        got["remove"] += len(t.get("remove") or [])
        got["create"] += 1 if t.get("status") == "create" else 0
        got["paths"].append(t.get("path"))
    out = [per[k] for k in order]
    for x in pv.get("excluded") or []:
        out.append({"chain": x.get("chain"), "process": x.get("process"),
                    "format": x.get("format"), "status": "will_not_run",
                    "copy": 0, "remove": 0, "create": 0, "paths": [],
                    "files": list(x.get("files") or []),
                    "reasons": list(x.get("reasons") or [])})
    return out


def build_plan_report(pv: dict, file_count: int = 0) -> str:
    """The staging PLAN as plain text — the confirmation dialog's View report,
    and the plan log.

    Same one-function rule as `build_report`: this text is what the window shows
    AND what the log file holds, so the two cannot describe different work. It
    carries the file NAMES — the dialog itself now shows counts only (the user's
    call: the delete list made the window too tall, and the summary line plus the
    per-row Delete column already say a delete is coming).
    """
    combos = plan_combinations(pv)
    bad = [c for c in combos if c["status"] == "will_not_run"]
    lines = [
        "OkGen — Run TOSCA Script: STAGING PLAN (nothing has run yet)",
        f"Script   : {pv.get('script')}",
        f"Selected : {file_count} file(s)",
        f"Plan     : {pv.get('copy_total', 0)} file(s) to copy into "
        f"{len(pv.get('targets') or [])} folder(s), "
        f"{pv.get('remove_total', 0)} existing file(s) to DELETE, "
        f"{len(bad)} combination(s) that will NOT run",
        "",
    ]
    if not pv.get("enabled", False):
        lines += ["STAGING OFF — input_staging.enabled is false in config/tosca.yaml,",
                  "              so the sheet would be updated but no files copied.", ""]
    elif not pv.get("configured", False):
        lines += ["NO INPUT FOLDERS — this script has no input_folders in "
                  "config/tosca.yaml,", "              so no files would be copied.", ""]
    lines.append("COMBINATIONS")
    if not combos:
        lines.append("  (none — nothing matched this script)")
    for c in combos:
        if c["status"] == "will_run":
            lines.append(
                f"  [WILL RUN] {str(c['chain'] or ''):<12} {str(c['process'] or ''):<15}"
                f" {str(c['format'] or ''):<26} copy {c['copy']}  delete {c['remove']}"
                + ("  (folder will be created)" if c["create"] else ""))
        else:
            lines.append(
                f"  [NOT RUN ] {str(c['chain'] or ''):<12} {str(c['process'] or ''):<15}"
                f" {str(c['format'] or ''):<26} {len(c['files'])} file(s)")
            for r in c.get("reasons") or []:
                lines.append(f"               reason : {r}")
            if c["files"]:
                lines.append("               files  :")
                lines += _wrap_names(c["files"], 24)
    # Per FOLDER, because that is where a delete actually happens — and the
    # names are here rather than in the dialog, so this is the only place they
    # can be checked before the run is agreed to.
    if pv.get("targets"):
        lines += ["", "PER FOLDER"]
        for t in pv["targets"]:
            lines.append(f"  {t.get('path')}")
            if t.get("status") == "create":
                lines.append("      (this folder does not exist yet and will be created)")
            if t.get("remove"):
                lines.append(f"      DELETE ({len(t['remove'])}):")
                lines += _wrap_names(list(t["remove"]), 8)
            else:
                lines.append("      DELETE (0): nothing to remove")
            lines.append(f"      COPY ({len(t.get('copy') or [])}):")
            lines += _wrap_names(list(t.get("copy") or []), 8)
    return "\n".join(lines) + "\n"


def log_folder(config) -> Path:
    """Where the run log goes: ``log_folder:`` in config/tosca.yaml, or the
    ``logs`` folder beside OkGen when that is blank or absent."""
    raw = str(((config.tosca() if config else {}) or {}).get("log_folder") or "").strip()
    return Path(raw) if raw else default_log_folder()


def _write_log(config, report: str, stamp: str, kind: str = "") -> Optional[str]:
    """Write the report beside OkGen. A log we cannot write must never fail the
    run — the report is still returned to the UI either way.

    ``kind`` names the file: ``okgen_tosca_<stamp>.log`` for a run,
    ``okgen_tosca_plan_<stamp>.log`` for a staging plan. Two names rather than
    one, because a plan and a run are different events and a folder where they
    are indistinguishable is a folder nobody reads twice.
    """
    try:
        folder = log_folder(config)
        fs.mkdir(folder, parents=True, exist_ok=True)
        suffix = f"{kind}_" if kind else ""
        path = folder / f"okgen_tosca_{suffix}{stamp}.log"
        fs.write_text(path, report, encoding="utf-8")
        return str(path)
    except (OSError, ValueError):
        return None


def log_plan(pv: dict, config, file_count: int = 0) -> dict:
    """Build the plan report and write it beside OkGen. Returns
    ``{"report", "log"}`` — ``log`` is None when it could not be written, which
    never fails anything: the text still reaches the window."""
    report = build_plan_report(pv, file_count)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return {"report": report, "log": _write_log(config, report, stamp, kind="plan")}


def run(paths, script_name, registry, config, launch=True, stage=True) -> dict:
    """Stage the selected files into the script's TOSCA input folders, populate
    its workbook from them, then fire its .bat (fire-and-forget) when rows were
    written — in that order (see the staging section above).

    ``launch=False`` skips the .bat and ``stage=False`` skips the file copy;
    both are for tests that don't want to spawn a process or touch a tree."""
    started_at = datetime.datetime.now()
    began = time.time()
    prep = _prepare(paths, script_name, registry, config)
    t = prep["t"]
    script = prep["script"]
    workbook = prep["workbook"]
    data_sheet = prep["data_sheet"]
    first_data_row = prep["first_data_row"]
    columns = prep["columns"]
    engines = prep["engines"]
    rows, errors, skipped = prep["rows"], prep["errors"], prep["skipped"]
    max_row = prep["max_row"]

    # STAGE FIRST — before the workbook is touched. TOSCA reads the files from
    # its own tree, so a run that updates the sheet without staging processes
    # whatever was left there last time. Doing it first also means a folder-level
    # failure (a lock, a permission) leaves the workbook untouched: there is
    # nothing to undo, and the sheet can never list a combination whose folder
    # was not set up. A combination whose folder is missing or misnamed is left
    # out of BOTH the staging and the sheet, and reported — the rest still runs.
    plan = plan_staging(rows, script, config, engines)
    staged = apply_staging(plan) if stage else {}
    rows = included_rows(rows, plan)

    max_clear_row = max(max_row, first_data_row + len(rows))
    if rows:
        try:
            write_data_sheet(workbook, data_sheet, rows, first_data_row, columns, max_clear_row)
        except PermissionError:
            # A TOSCA run stopped part-way leaves the workbook open in Excel, and
            # the lock outlives it by however long that Excel sits there. Ask
            # Excel to close THAT workbook (never kill Excel), then try once
            # more. A workbook with unsaved changes is left alone — see
            # close_open_workbook.
            outcome = (close_open_workbook(workbook)
                       if bool(t.get("close_open_workbook", True)) else "disabled")
            if outcome != CLOSE_OK:
                raise ToscaError(_locked_message(workbook, outcome))
            try:
                write_data_sheet(workbook, data_sheet, rows, first_data_row,
                                 columns, max_clear_row)
            except PermissionError:
                raise ToscaError(_locked_message(workbook, CLOSE_OK))
            except OSError as exc:
                raise ToscaError(f"could not write the workbook {workbook}: {exc}")
        except OSError as exc:
            raise ToscaError(f"could not write the workbook {workbook}: {exc}")

    # Fire the selected script's .bat — only when we actually wrote rows (nothing
    # to run otherwise). Fire-and-forget: launch detached and return at once.
    launched = False
    launch_error = None
    bat_cfg = script.get("bat")
    bat_file = None
    if rows and launch:
        if not bat_cfg:
            launch_error = "no .bat configured for this script (sheet updated only)"
        else:
            try:
                bat_file = _resolve_one(bat_cfg, _BAT_EXTS, "TOSCA .bat")
                _launch_bat(bat_file)
                launched = True
            except ToscaError as exc:
                launch_error = str(exc)
            except Exception as exc:                    # noqa: BLE001
                launch_error = f"failed to start .bat: {exc}"

    res = {
        "workbook": str(workbook),
        "script": script_name,
        "data_sheet": data_sheet,
        "started": started_at.strftime("%Y-%m-%d %H:%M:%S"),
        "elapsed_seconds": round(time.time() - began, 1),
        "written": len(rows),
        "rows": rows,
        "errors": errors,
        "skipped": skipped,
        "staging": {
            "enabled": plan.get("enabled", False),
            "configured": plan.get("configured", False),
            "folders": staged.get("folders", []),
            "removed": staged.get("removed", 0),
            "copied": staged.get("copied", 0),
            "created": staged.get("created", 0),
            "excluded": plan.get("excluded", []),
        },
        "applies_to": sorted(engines),
        "launched": launched,
        "bat": str(bat_file) if bat_file else bat_cfg,
        "launch_error": launch_error,
    }
    # The per-combination roll-up the summary window shows, computed once here so
    # the window and the report cannot disagree about a count.
    res["combinations"] = combinations(res)
    # EVERY run is logged, not only a failing one (the user's call, and what Send
    # already does): "what did last Tuesday's run actually stage?" is the
    # question a log exists to answer, and by then the window is long gone.
    res["report"] = build_report(res)
    res["log"] = _write_log(config, res["report"],
                            started_at.strftime("%Y%m%d_%H%M%S"))
    return res
