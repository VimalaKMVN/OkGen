"""Compile xlsx layout-definition files into :class:`Layout` objects.

Each xlsx tab is laid out as:

    row N   : a real sample record (marker already stripped)
    row N+1 : column headers (field_name, field_size, output_field_name,
              Position, Value, field_type, field_name_section, ...)
    row N+2+: one row per field definition

Tabs without a recognizable header row (e.g. free-form notes sheets) are
skipped. ``Position`` in the spec is formula-computed and frequently carries
``#VALUE!`` / ``NULL`` errors, so we recompute ``start`` from the cumulative
sum of ``field_size`` and keep the spec's ``Position`` only as
``declared_position`` for cross-checking.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from okgen.layout.models import Field, Layout, Section

# Spreadsheet sentinels that mean "no usable value".
_BAD_VALUES = {"", "NULL", "#VALUE!", "#REF!", "#N/A", "NONE"}

# Header cells we use to recognize a field-definition table within a tab.
_REQUIRED_HEADERS = {"field_name", "output_field_name", "field_size"}


def _norm(value) -> str:
    """Normalize a cell to a stripped string ('' for None)."""
    if value is None:
        return ""
    return str(value).strip()


def _clean_int(value) -> Optional[int]:
    """Parse an int from a messy cell; return None for blanks/errors."""
    s = _norm(value)
    if s.upper() in _BAD_VALUES:
        return None
    # Excel may hand back floats like "2.0".
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


# Filename prefixes whose layouts use a delimited (not fixed-width) record
# format. EU/EWMS pretickets are pipe-delimited; NA layouts are fixed-width.
_DELIMITED_PREFIXES = ("TJXEWMS_",)


def _layout_name_from_filename(filename: str) -> str:
    """'TJXNA_CartonLabelLayout.xlsx' -> 'CartonLabel';
    'TJXEWMS_PreticketLayout.xlsx' -> 'EUPreticket'."""
    stem = Path(filename).stem
    stem = re.sub(r"^TJXNA_", "", stem)
    stem = re.sub(r"^TJXEWMS_", "EU", stem)   # EU/EWMS layouts keep an EU prefix
    stem = re.sub(r"Layout$", "", stem)
    return stem or stem


def _is_delimited_filename(filename: str) -> bool:
    return Path(filename).name.startswith(_DELIMITED_PREFIXES)


def _find_header_row(rows: List[List]) -> Optional[int]:
    """Index of the row that looks like the column-header row, or None."""
    for i, row in enumerate(rows):
        cells = {_norm(c).lower() for c in row}
        if _REQUIRED_HEADERS.issubset(cells):
            return i
    return None


def _column_index(header_row: List, *names: str) -> Optional[int]:
    """First column index whose header matches any of ``names`` (case-insensitive)."""
    lowered = [_norm(c).lower() for c in header_row]
    for name in names:
        target = name.lower()
        for idx, cell in enumerate(lowered):
            if cell == target:
                return idx
    return None


def _compile_section(tab_title: str, rows: List[List]) -> Optional[Section]:
    """Build a Section from a worksheet's rows, or None if it isn't a layout tab."""
    hdr_idx = _find_header_row(rows)
    if hdr_idx is None:
        return None

    header = rows[hdr_idx]
    col = {
        "section": _column_index(header, "field_name_section"),
        "field_id": _column_index(header, "field_id"),
        "field_name": _column_index(header, "field_name"),
        "size": _column_index(header, "field_size"),
        "out_name": _column_index(header, "output_field_name"),
        "type": _column_index(header, "field_type"),
        "position": _column_index(header, "Position"),
        "value": _column_index(header, "Value"),
    }

    def cell(row, key):
        idx = col[key]
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    section_name = tab_title.strip()
    fields: List[Field] = []
    for row in rows[hdr_idx + 1:]:
        out_name = _norm(cell(row, "out_name"))
        field_name = _norm(cell(row, "field_name"))
        if not out_name and not field_name:
            continue  # blank/spacer row

        sec = _norm(cell(row, "section"))
        if sec:
            section_name = sec  # prefer the spec's section label

        size = _clean_int(cell(row, "size"))
        declared = _clean_int(cell(row, "position"))
        raw_value = cell(row, "value")
        sample_value = None if raw_value is None else str(raw_value)
        ftype = _norm(cell(row, "type")) or "char"

        f = Field(
            name=out_name or field_name,
            field_name=field_name,
            size=size,
            start=None,  # filled in below
            field_type=ftype,
            declared_position=declared,
            sample_value=sample_value,
            field_id=_norm(cell(row, "field_id")) or None,
        )
        if size is None:
            f.issues.append("missing/invalid field_size in spec")
        fields.append(f)

    # Compute positions over the full field sequence first (so a sized field
    # following an unsized one can still resync), then drop the unsized rows.
    # Those are spec artifacts: repeating-record extras (lane2..lane10, where
    # lane1's position/size applies to every lane record) and computed
    # aggregates (*_totqty). They are recorded in `ignored_fields`.
    _recompute_positions(fields)
    sized = [f for f in fields if f.size is not None]
    ignored = [f.name for f in fields if f.size is None]

    # The sample record lives in column A somewhere above the header row
    # (often row 0, with blank spacer rows in between).
    sample_record = None
    for prior in rows[:hdr_idx]:
        if prior and prior[0] is not None and str(prior[0]).strip():
            sample_record = str(prior[0])
            break
    section = Section(
        name=section_name, tab=tab_title, fields=sized,
        sample_record=sample_record, ignored_fields=ignored,
    )
    section.record_length = max(
        (f.end for f in sized if f.end is not None), default=None
    )
    _flag_section_issues(section)
    return section


def _recompute_positions(fields: List[Field]) -> None:
    """Assign 1-based start from cumulative sizes; resync on a valid Position."""
    cursor: Optional[int] = 1
    for f in fields:
        if cursor is None and f.declared_position is not None:
            # Recover the running offset from the spec's Position column.
            cursor = f.declared_position
            f.issues.append("start resynced from declared Position")
        f.start = cursor
        if cursor is not None and f.size is not None:
            cursor = cursor + f.size
        else:
            cursor = None  # unknown size breaks the running offset


def _flag_section_issues(section: Section) -> None:
    """Record overlaps, gaps, and Position/recompute mismatches."""
    prev_end = 0
    for f in section.fields:
        if f.start is None or f.size is None:
            continue
        if f.start <= prev_end:
            section.issues.append(
                f"{f.name}: overlaps previous field (start {f.start} <= {prev_end})"
            )
        elif f.start > prev_end + 1:
            section.issues.append(
                f"{f.name}: gap before field (start {f.start}, prev end {prev_end})"
            )
        if f.declared_position is not None and f.declared_position != f.start:
            f.issues.append(
                f"declared Position {f.declared_position} != computed start {f.start}"
            )
        prev_end = f.end


def compile_layout(xlsx_path: Path) -> Layout:
    """Compile one xlsx definition file into a :class:`Layout`."""
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        layout = Layout(
            name=_layout_name_from_filename(xlsx_path.name),
            source_file=xlsx_path.name,
            delimited=_is_delimited_filename(xlsx_path.name),
        )
        for ws in wb.worksheets:
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            section = _compile_section(ws.title, rows)
            if section is None:
                continue
            if layout.ticket_process is None:
                # Pull ticket_process from the first field row that carries it.
                tp = _ticket_process(rows)
                layout.ticket_process = tp
            layout.sections.append(section)
        if not layout.sections:
            layout.issues.append("no layout tabs found")
        _derive_section_markers(layout, xlsx_path.parent)
        return layout
    finally:
        wb.close()


# Leading chars treated as a one-char record marker on fixed-width detail lines.
# (Kept in sync with okfile.DETAIL_MARKERS; duplicated here to avoid an import
# cycle between the compiler and the parser.)
_DETAIL_MARKERS = set("|#&")


def _derive_section_markers(layout: Layout, data_dir: Path) -> None:
    """Assign each non-header section its canonical record-type ``marker`` and a
    ``sample_raw`` seed line.

    The marker is what lets records route to the *right* section even when an
    earlier section is empty (e.g. no Lane records): the parser matches each
    line's marker to the section that owns it instead of mapping the k-th marker
    seen to the k-th section. ``sample_raw`` is a genuine full record line used
    to seed the first row when a user adds records to an empty section.

    * Delimited layouts carry the marker in the section's sample record
      (``|&|...`` / ``|#|...``), so it is read straight from ``sample_record``.
    * Fixed-width layouts strip the marker from their xlsx sample, so it is
      learned from the sibling reference ``<name>.OK`` file (a complete, in-order
      example) — which is also where ``sample_raw`` comes from for every layout.

    If that reference file is absent/unreadable the markers stay None (the
    parser falls back to its original order-of-appearance rule) and no seed is
    stored (adding to a still-empty section is then refused with a clear error).
    """
    detail_secs = layout.sections[1:]
    if not detail_secs:
        return
    if layout.delimited:
        for sec in detail_secs:
            sec.marker = _delim_marker(sec.sample_record or "", layout.delimiter)

    ref = data_dir / f"{layout.name}.OK"
    try:
        lines = ref.read_bytes().decode("latin-1").splitlines()
    except OSError:
        return

    # First real line seen per record-type token, in order of appearance, plus
    # the first ALL-BLANK (zero/space) line per token — some formats pad their
    # detail block with such filler rows (Preticket), and the fill feature reuses
    # them verbatim rather than synthesising (the zero/space byte layout is not
    # derivable from the field spec).
    first_line: dict = {}
    filler_line: dict = {}
    for raw in lines[1:]:                       # line 0 is the header
        if layout.delimited:
            token = _delim_marker(raw, layout.delimiter)
        else:
            first = raw[:1]
            token = first if first in _DETAIL_MARKERS else ""
        if _is_blank_line(raw):
            filler_line.setdefault(token, raw)
        else:
            first_line.setdefault(token, raw)

    if layout.delimited:
        # Markers already set from the xlsx samples; match tokens by marker.
        by_marker = {sec.marker: sec for sec in detail_secs}
        for token, raw in first_line.items():
            sec = by_marker.get(token)
            if sec is not None and sec.sample_raw is None:
                sec.sample_raw = raw
        return

    # Fixed-width: map tokens to detail sections in appearance order (marker-less
    # "" tokens always go to the first detail section, like the parser).
    markered = 0
    for token, raw in first_line.items():
        if token == "":
            sec = detail_secs[0]
        else:
            sec = detail_secs[markered] if markered < len(detail_secs) else None
            markered += 1
        if sec is None:
            continue
        if sec.marker is None:
            sec.marker = token
        if sec.sample_raw is None:
            sec.sample_raw = raw
        if sec.filler_raw is None and token in filler_line:
            sec.filler_raw = filler_line[token]


def _is_blank_line(raw: str) -> bool:
    """True when a record line is an all-zero/space filler row: its content is
    only '0' and ' ', ignoring the trailing terminator ('\\') and CR/LF/pad."""
    core = raw.rstrip("\r\n").rstrip(" ").rstrip("\\").rstrip(" ")
    return core != "" and set(core) <= set("0 ")


def _delim_marker(raw: str, delimiter: str) -> str:
    """Record-type marker of a delimited sample: the char after a leading
    delimiter (``|&|`` -> ``&``), or '' for marker-less delimited details."""
    if raw[:1] == delimiter:
        return raw[1:2]
    return ""


def _ticket_process(rows: List[List]) -> Optional[str]:
    hdr_idx = _find_header_row(rows)
    if hdr_idx is None:
        return None
    idx = _column_index(rows[hdr_idx], "ticket_process")
    if idx is None:
        return None
    for row in rows[hdr_idx + 1:]:
        if idx < len(row):
            val = _norm(row[idx])
            if val and val.upper() not in _BAD_VALUES:
                return val
    return None


def compile_dir(data_dir: Path) -> List[Layout]:
    """Compile every ``*.xlsx`` in ``data_dir`` into Layouts."""
    data_dir = Path(data_dir)
    layouts = []
    for xlsx in sorted(data_dir.glob("*.xlsx")):
        layouts.append(compile_layout(xlsx))
    return layouts
